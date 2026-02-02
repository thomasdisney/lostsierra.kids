#!/usr/bin/env python3
"""
Jira Campaign Update Tool
Automatically processes Jira campaign exports and generates formatted Excel reports.
"""

import json
import sys
import os
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

# Set UTF-8 encoding for console output on Windows
if sys.platform == "win32":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


def get_downloads_folder():
    """Get the user's Downloads folder path."""
    return Path.home() / "Downloads"


def find_latest_jira_export():
    """Find the most recent file in Downloads containing 'jira'."""
    downloads = get_downloads_folder()

    # Find all files containing 'jira' (case-insensitive)
    jira_files = [
        f for f in downloads.glob("*")
        if f.is_file() and "jira" in f.name.lower()
    ]

    if not jira_files:
        print("[ERROR] No Jira export file found in Downloads folder")
        print("[ERROR] Please export your campaigns from Jira and save to Downloads")
        sys.exit(1)

    # Sort by modification time (most recent first)
    latest = sorted(jira_files, key=lambda f: f.stat().st_mtime, reverse=True)[0]

    # Validate file is recent (not older than current date)
    file_mtime = datetime.fromtimestamp(latest.stat().st_mtime)
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    if file_mtime.date() < today.date():
        print(f"[ERROR] Jira export file is stale (last modified: {file_mtime.date()})")
        print(f"[ERROR] Please export a fresh copy from Jira")
        sys.exit(1)

    return latest


def load_config(config_path):
    """Load campaign definitions from JSON config file."""
    if not config_path.exists():
        # Create template config if it doesn't exist
        template = {
            "campaigns": [
                {
                    "id": "PROJ-1",
                    "name": "Example Campaign",
                    "description": "Campaign description",
                    "owner": "Team Name",
                    "custom_fields": {}
                }
            ],
            "jira_base_url": "https://yourcompany.atlassian.net",
            "output_settings": {
                "output_folder": "./reports",
                "include_fields": ["key", "summary", "status", "assignee", "priority", "labels"]
            }
        }

        with open(config_path, "w") as f:
            json.dump(template, f, indent=2)

        print(f"[OK] Created template config: {config_path}")
        return template

    try:
        with open(config_path, "r") as f:
            config = json.load(f)
        print(f"[OK] Loaded campaign definitions from {config_path.name}")
        return config
    except json.JSONDecodeError as e:
        print(f"[ERROR] Error parsing config file: {e}")
        sys.exit(1)


def read_jira_export(file_path):
    """Read Jira export file (Excel or CSV)."""
    try:
        if file_path.suffix.lower() in ['.csv']:
            df = pd.read_csv(file_path)
        else:
            # Try to read as Excel, specify engine if needed
            try:
                df = pd.read_excel(file_path)
            except:
                # Fallback to openpyxl engine
                df = pd.read_excel(file_path, engine='openpyxl')

        print(f"[OK] Read Jira export: {file_path.name}")
        print(f"  - {len(df)} total tickets")
        return df
    except Exception as e:
        print(f"[ERROR] Error reading file: {e}")
        sys.exit(1)


def process_campaigns(df):
    """Filter and process campaign data from the export."""
    # Make column names case-insensitive for lookup
    df_cols = {col.lower(): col for col in df.columns}

    # Try to find the issue type column
    issue_type_col = None
    for potential_col in ["issue type", "issuetype", "type"]:
        if potential_col in df_cols:
            issue_type_col = df_cols[potential_col]
            break

    # Filter for campaigns
    if issue_type_col:
        campaigns_df = df[df[issue_type_col].str.lower() == "campaign"].copy()
    else:
        # If no issue type column, assume all rows are campaigns
        campaigns_df = df.copy()

    print(f"[OK] Found {len(campaigns_df)} campaign tickets")

    return campaigns_df, df


def add_autofilter(worksheet, df):
    """Add AutoFilter dropdown to worksheet headers."""
    if len(df) == 0:
        return

    # Get the range for the data
    max_row = len(df) + 1
    max_col = len(df.columns)

    # Convert to Excel column letters
    from openpyxl.utils import get_column_letter
    end_col = get_column_letter(max_col)

    # Add AutoFilter
    worksheet.auto_filter.ref = f"A1:{end_col}{max_row}"


def create_formatted_workbook(campaigns_df, all_df, config, output_path):
    """Create a formatted Excel workbook with multiple sheets."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define colors for status
    status_colors = {
        "done": "C6EFCE",
        "in progress": "FFF2CC",
        "todo": "FCE4D6",
        "blocked": "F8CBAD",
    }

    # Define styles
    header_fill = PatternFill(start_color="1e3a2f", end_color="1e3a2f", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Create Summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    summary_data = [
        ["Campaign Update Report", ""],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["Total Campaigns", len(campaigns_df)],
        ["Total Tickets", len(all_df)],
    ]

    # Add status breakdown if status column exists
    if any(col.lower() == "status" for col in campaigns_df.columns):
        status_col = next(col for col in campaigns_df.columns if col.lower() == "status")
        status_counts = campaigns_df[status_col].value_counts()
        summary_data.append(["", ""])
        summary_data.append(["Campaign Status Breakdown", ""])
        for status, count in status_counts.items():
            summary_data.append([status, count])

    for row_idx, row in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row, 1):
            cell = ws_summary.cell(row_idx, col_idx, value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15

    # Create Raw Data sheet (consolidated source)
    ws_raw = wb.create_sheet("Raw Data", 1)

    # Select relevant fields only
    relevant_fields = [col for col in all_df.columns
                      if col.lower() not in ["project id", "project url"]]
    raw_display = all_df[relevant_fields].copy()

    # Rename columns for display
    column_mapping = {col: col.replace("_", " ").title() for col in raw_display.columns}
    raw_display = raw_display.rename(columns=column_mapping)

    # Write headers
    for col_idx, column_title in enumerate(raw_display.columns, 1):
        cell = ws_raw.cell(1, col_idx, column_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write data rows
    for row_idx, row in enumerate(dataframe_to_rows(raw_display, index=False, header=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws_raw.cell(row_idx, col_idx, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Auto-fit columns
    for column in ws_raw.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_raw.column_dimensions[column_letter].width = adjusted_width

    # Add AutoFilter to raw data
    ws_raw.freeze_panes = "A2"
    add_autofilter(ws_raw, raw_display)

    # Create Campaign Details sheet
    ws_details = wb.create_sheet("Campaign Details", 2)

    # Prepare data for details sheet
    display_df = campaigns_df.copy()

    # Rename columns for display
    column_mapping = {col: col.replace("_", " ").title() for col in display_df.columns}
    display_df = display_df.rename(columns=column_mapping)

    # Write headers
    for col_idx, column_title in enumerate(display_df.columns, 1):
        cell = ws_details.cell(1, col_idx, column_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Write data rows
    for row_idx, row in enumerate(dataframe_to_rows(display_df, index=False, header=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws_details.cell(row_idx, col_idx, value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Apply status colors
            if col_idx < len(row) and "status" in display_df.columns[col_idx - 1].lower():
                status_key = str(value).lower() if value else ""
                for key, color in status_colors.items():
                    if key in status_key:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        break

    # Auto-fit columns
    for column in ws_details.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_details.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row and add AutoFilter
    ws_details.freeze_panes = "A2"
    add_autofilter(ws_details, display_df)

    # Create Timeline sheet if date columns exist
    date_columns = [col for col in display_df.columns if "date" in col.lower() or "created" in col.lower()]
    if date_columns:
        ws_timeline = wb.create_sheet("Timeline", 3)

        # Sort by first date column found
        timeline_df = display_df.copy()
        try:
            timeline_df = timeline_df.sort_values(date_columns[0])
        except:
            pass

        # Write headers
        for col_idx, column_title in enumerate(timeline_df.columns, 1):
            cell = ws_timeline.cell(1, col_idx, column_title)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Write data rows
        for row_idx, row in enumerate(dataframe_to_rows(timeline_df, index=False, header=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_timeline.cell(row_idx, col_idx, value)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Auto-fit columns
        for column in ws_timeline.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_timeline.column_dimensions[column_letter].width = adjusted_width

        # Add AutoFilter
        ws_timeline.freeze_panes = "A2"
        add_autofilter(ws_timeline, timeline_df)

    # Save workbook
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"[OK] Generated report: {output_path}")


def main():
    """Main execution flow."""
    script_dir = Path(__file__).parent
    config_path = script_dir / "campaign_definitions.json"

    print("[JIRA] Campaign Update Tool\n")

    # Load configuration
    config = load_config(config_path)

    # Find latest Jira export (with validation)
    jira_file = find_latest_jira_export()
    if not jira_file:
        sys.exit(1)

    print(f"[OK] Found: {jira_file.name}\n")

    # Read and process data
    all_df = read_jira_export(jira_file)
    campaigns_df, all_df = process_campaigns(all_df)

    if len(campaigns_df) == 0:
        print("[WARN] No campaigns found in export")
        sys.exit(0)

    # Generate output
    output_folder = script_dir / config.get("output_settings", {}).get("output_folder", "reports")
    timestamp = datetime.now().strftime("%Y-%m-%d")
    output_file = output_folder / f"campaign_update_{timestamp}.xlsx"

    create_formatted_workbook(campaigns_df, all_df, config, output_file)

    print(f"\n[SUCCESS] Report saved to: {output_file}")

    # Open the file with default application
    try:
        if sys.platform == "win32":
            os.startfile(output_file)
        elif sys.platform == "darwin":
            os.system(f'open "{output_file}"')
        else:
            os.system(f'xdg-open "{output_file}"')
    except Exception as e:
        print(f"[INFO] Could not auto-open file: {e}")


if __name__ == "__main__":
    main()
